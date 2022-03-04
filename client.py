"""
written by : 
  *******************************************************************************
                    __                             __    __               _ 
   ____ ___  ____  / /_  ________  ____     ____ _/ /_  / /_  ____ ______(_)
  / __ `__ \/ __ \/ __ \/ ___/ _ \/ __ \   / __ `/ __ \/ __ \/ __ `/ ___/ / 
 / / / / / / /_/ / / / (__  )  __/ / / /  / /_/ / /_/ / /_/ / /_/ (__  ) /  
/_/ /_/ /_/\____/_/ /_/____/\___/_/ /_/   \__,_/_.___/_.___/\__,_/____/_/   
                                                                            
name of project : communication-four-client-with-server-and-sending-data-
                    from-exel-file-(.xlsx)-from-server-to-clients
*********************************************************************************
"""
from colorama import Fore,init
import socket
import os
import time
import sys
import threading
import random
import re

init()
try:
    from openpyxl import load_workbook,Workbook
    from openpyxl.utils import get_column_letter
except:
    os.system('pip instal openpyxl')    


# variable
ip=''
port=0
server=None
connection=None

logo="""

 ██████╗██╗     ██╗███████╗███╗   ██╗████████╗
██╔════╝██║     ██║██╔════╝████╗  ██║╚══██╔══╝
██║     ██║     ██║█████╗  ██╔██╗ ██║   ██║   
██║     ██║     ██║██╔══╝  ██║╚██╗██║   ██║   
╚██████╗███████╗██║███████╗██║ ╚████║   ██║   
 ╚═════╝╚══════╝╚═╝╚══════╝╚═╝  ╚═══╝   ╚═╝   
                                              
"""
print(Fore.RED+logo)
time.sleep(0.1)
print(Fore.CYAN+'-----------------------------------------------')
time.sleep(0.1)
while True:
    octed_status=False
    try:
        print(Fore.YELLOW+'')
        ip = input("┌─["+"ENTER IP OF SERVER"+"""]
└──╼ """+"卐 ")
        if ip==None or ip=="" or ip=="\n":
            print(Fore.RED+'the ip is empty'.upper())
            continue
        condition=str(ip).split('.')
        for dot in condition:
            if not dot.isdigit():
                octed_status=True
                break
        if octed_status:
            print(Fore.RED+'one of the octed is not correct'.upper())
            continue


        if len(condition)!=4:
            print(Fore.RED+'your ip is not correct '.upper())
            continue

        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the ip'.upper())    


# print(ip)
time.sleep(0.1)
while True:
    try:
        print(Fore.YELLOW+'')
        port = input("┌─["+"ENTER PORT OF SERVER"+"""]
└──╼ """+"卐 ")
        if port==None or port=="" or port=="\n" or port=='0':
            print(Fore.RED+'the port is empty'.upper())
            continue
        if not port.isdigit():
            print(Fore.RED+'number of port is not number'.upper())
            continue
        port=int(port)
        if port<1000:
            print(Fore.RED+'number of port has to be more 1000 '.upper())
            continue
        
        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the ip'.upper())  

try:

    connection=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
    connection.connect((ip,port))
    print('clinet with ip ',str(connection.getsockname()[0]),' connected to server'.upper())
except:
    print('i can not connect to server '.upper())
# check all client online
status=False
while True:
    if not status: 
        input('press enter to connect ...'.upper())
        connection.send('wait'.encode())
        ack=connection.recv(1234).decode()
        if ack=='fail':
            print('all of client have to be online ...'.upper())
            continue
    status=True
    ack=connection.recv(1234).decode()
    print(ack)
    break

while True:
    try:
        data=connection.recv(1234).decode()
        # print(data)
        sp=data.split('-----') 
        # print(len(sp))
        wb=Workbook()
        # remove all default sheet
        for sheet in wb.sheetnames:
            ws=wb[sheet]
            wb.remove(ws)
        # set name for sheet
        for i in range(len(sp)):
            wb.create_sheet('sheet'+str(i+1))
        # get correct data 
        dic_ex={}
        for index,item in enumerate(sp):
            list_=[]
            xx=re.findall(r'[0-9,]*[0-9]*',item)
            for x in xx:
                if len(x)!=0:
                    list_.append(x)
            dic_ex['sheet'+str(index+1)]=list_
        # print(dic_ex)   

        # set data to exel
        for key,value in dic_ex.items():
            ws=wb[key]
            for val in value:
                val=str(val).split(',')
                ws.append(val)
        wb.save(f'{connection.getsockname()[0]}.xlsx')
        print(Fore.GREEN+'seccussful to recaive data'.upper())
        connection.send('true'.encode())
    except:
        print(Fore.RED+'fail to recaive data'.upper())
        connection.send('false'.encode())


                