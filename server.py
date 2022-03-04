"""
written by : 
  *******************************************************************************
                    __                             __    __               _ 
   ____ ___  ____  / /_  ________  ____     ____ _/ /_  / /_  ____ ______(_)
  / __ `__ \/ __ \/ __ \/ ___/ _ \/ __ \   / __ `/ __ \/ __ \/ __ `/ ___/ / 
 / / / / / / /_/ / / / (__  )  __/ / / /  / /_/ / /_/ / /_/ / /_/ (__  ) /  
/_/ /_/ /_/\____/_/ /_/____/\___/_/ /_/   \__,_/_.___/_.___/\__,_/____/_/   
                                                                            
name of project : communication-four-client-with-server-and-sending-data-
                    from-exel-file-(.xlsx)-from-server-to-clients
*********************************************************************************
"""
import socket
import os
import time
import sys
import threading
import random
from queue import Queue
import matplotlib.pyplot as plt


try:
    from openpyxl import Workbook,load_workbook
    from openpyxl.utils import get_column_letter
except:
    os.system('pip install openpyxl')    

try:
    from colorama import Fore,init
    init()
except:
    os.system('pip install colorama')    

try:
    import networkx as nx
except:
    os.system('pip install networkx')        




# variable
ip=''
port=0
server=None
client=None
addr=''

all_connection=[]
menu="""
[b] BROADCASRT
[u] UNICAST
[g] GRAPH
"""

logo="""

███████╗███████╗██████╗ ██╗   ██╗███████╗██████╗ 
██╔════╝██╔════╝██╔══██╗██║   ██║██╔════╝██╔══██╗
███████╗█████╗  ██████╔╝██║   ██║█████╗  ██████╔╝
╚════██║██╔══╝  ██╔══██╗╚██╗ ██╔╝██╔══╝  ██╔══██╗
███████║███████╗██║  ██║ ╚████╔╝ ███████╗██║  ██║
╚══════╝╚══════╝╚═╝  ╚═╝  ╚═══╝  ╚══════╝╚═╝  ╚═╝
                                                 
"""
print(Fore.RED+logo)
time.sleep(0.1)
print('-----------------------------------------------')
time.sleep(0.1)


while True:
    octed_status=False
    try:
        print(Fore.YELLOW+'')
        ip = input("┌─["+"ENTER IP OF SERVER"+"""]
└──╼ """+"卐 ")
        if ip==None or ip=="" or ip=="\n":
            print(Fore.RED+'the ip is empty'.upper())
            continue
        # 192.168.1.200
        condition=str(ip).split('.')
        for dot in condition:
            if not dot.isdigit():
                octed_status=True
                break
        if octed_status:
            print(Fore.RED+'one of the octed is not correct'.upper())
            continue


        if len(condition)!=4:
            print(Fore.RED+'your ip is not correct '.upper())
            continue

        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the ip'.upper())    


# print(ip)
time.sleep(0.1)
while True:
    try:
        print(Fore.YELLOW+'')
        port = input("┌─["+"ENTER PORT OF SERVER"+"""]
└──╼ """+"卐 ")
       
        if not port.isdigit():
            print(Fore.RED+'number of port is not number'.upper())
            continue
        port=int(port)

        if port<1000:
            print(Fore.RED+'number of port has to be more 1000'.upper())
            continue

        
        break

    except KeyboardInterrupt:
        sys.exit()    
    except:
        print(Fore.RED+'i cant get the port'.upper())  

print('-----------------------------------------------')
time.sleep(0.1)

 
try:
    server=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
    server.bind((ip,int(port)))
    server.listen(4)
    print(Fore.GREEN+'server with ip  : ',server.getsockname()[0],' is running on port : '.upper(),port)
except KeyboardInterrupt:
    sys.exit()    
except:
    print(Fore.RED+'i cant bind the server'.upper())  

def show_xlsx_file():
    # show all xlsx file ... 
    xlsx_list=[]
    all_file=os.listdir(os.getcwd())
    for file in all_file:
        sp =file.split('.')
        if sp[1]=='xlsx':
            xlsx_list.append(file)

     
    while True:    
        for index , file_xlsx in enumerate(xlsx_list,start=1):
            print(Fore.YELLOW+f'[ {index} ] {file_xlsx}') 
        i=input(Fore.GREEN+'choose a number >>>> '.upper())
        if i==None or i=='' or i==' ' or i=='\n':
            print(Fore.RED+'empty data !!!'.upper())
            show_xlsx_file()
        if not i.isdigit():
            print(Fore.RED+'not number !!!'.upper())
            show_xlsx_file()
        i=int(i)
        break
    print(Fore.BLUE+f'{xlsx_list[i-1]} choosed ... ')  
    # load data from xlsx file ...
    wb=load_workbook(xlsx_list[i-1]) 
    data=''
    for _index,sheet in enumerate( wb.sheetnames):
        var=''
        ws=wb[sheet]
        for row in range(1,ws.max_row+1):
            for col in range(1,ws.max_column+1):
                char =get_column_letter(col)
                value=ws[char+str(row)].value
                if col!=ws.max_column:
                    var+=str(value)+','
                else:
                    var+=str(value)+'\n' 
        if _index!=len(wb.sheetnames)-1:                 
            data=data+var+'-----'+'\n'
        else:
            data=data+var    

    wb.save(xlsx_list[i-1]) 
    return data
 



ip_dic={}
def check_all_client_online(client,ip):
    # check all client online 
    while True:
        client.recv(1234).decode()
        if ip not in ip_dic.keys():
                ip_dic[ip]=client      
        if len(ip_dic)!=2:
            print(Fore.RED+'all of client have to be online...'.upper()) 
            for cli in ip_dic.values():
                cli.send('fail'.encode())    
        else:
            print(Fore.GREEN+'connection seccuessful...'.upper())  
            for cli in ip_dic.values():
                cli.send('connection seccuesful'.encode())
            break    

    time.sleep(1)

    while True:
        time.sleep(0.1)
        print(Fore.CYAN+menu)
        brod_uni=input(Fore.YELLOW+'enter one of mode : '.upper()).lower()
        if brod_uni=='b':
            data=show_xlsx_file()
            for cli in ip_dic.values():
                cli.send(data.encode())
            for ip,cli in ip_dic.items():
                message=cli.recv(12345).decode()
                if message=='true':
                    print(f'{ip} says seccusful to recaive data'.upper()) 
                elif message=='false':
                    print(f'{ip} says fail to recaive data'.upper()) 
                   
        elif brod_uni=='u':
            data=show_xlsx_file()
            host=input(Fore.YELLOW+'enter ip of client : '.upper())
            if host in ip_dic.keys():   
                ip_dic[host].send(data.encode())
                message=ip_dic[host].recv(12345).decode()
                if message=='true':
                        print(f'{host} says seccusful to recaive data'.upper()) 
                elif message=='false':
                        print(f'{host} says fail to recaive data'.upper())
            else:
                print(Fore.RED+'there is not client with ip : '.upper(),host)    
             

        elif brod_uni=='g':
            g=nx.Graph()
            g.add_node('server')
            for ip in ip_dic.keys():
                g.add_node(ip)
                g.add_edge(ip,'server')    
            nx.draw(g,with_labels=True, font_weight='bold',node_size =[8000])
            plt.show()


        else:
            print(Fore.RED+'you have to use b or u or g')
            continue    


while True:
    try:
        client,addr=server.accept()
        threading._start_new_thread(check_all_client_online,(client,addr[0]))
          
    except:
        print('no client cant connect to me')    